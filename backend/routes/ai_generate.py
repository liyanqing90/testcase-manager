from flask import Blueprint, request, jsonify, send_file
import sys
import os
import asyncio
import subprocess
import uuid
import json
from datetime import datetime
from werkzeug.utils import secure_filename
import shutil
from backend.models.db import execute_query

# 添加AI测试系统路径
sys.path.append(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'ai_test_cases', 'src'))

ai_generate_bp = Blueprint('ai_generate', __name__)

# 配置上传目录
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'ai_test_cases', 'docs')
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'md', 'txt'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@ai_generate_bp.route('/upload', methods=['POST'])
def upload_document():
    """上传需求文档"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件上传'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400

        if file and allowed_file(file.filename):
            # 获取原始文件名和扩展名
            original_filename = file.filename
            name, ext = os.path.splitext(original_filename)

            # 生成安全的文件名，但保留扩展名
            safe_name = secure_filename(name)
            if not safe_name:  # 如果安全文件名为空，使用默认名称
                safe_name = "document"

            # 添加时间戳避免重名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{safe_name}{ext}"

            # 确保上传目录存在
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)

            # 保存文件
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(file_path)

            return jsonify({
                'success': True,
                'filename': filename,
                'file_path': file_path,
                'message': '文件上传成功'
            })
        else:
            return jsonify({'error': '不支持的文件格式'}), 400

    except Exception as e:
        return jsonify({'error': f'上传失败: {str(e)}'}), 500


@ai_generate_bp.route('/generate', methods=['POST'])
def generate_test_cases():
    """生成测试用例"""
    try:
        data = request.json
        filename = data.get('filename')
        output_filename = data.get('output_filename', 'test_cases.xlsx')
        test_type = data.get('test_type', 'functional')
        concurrency = data.get('concurrency', 1)

        if not filename:
            return jsonify({'error': '文件名不能为空'}), 400

        # 构建文件路径
        doc_path = os.path.join(UPLOAD_FOLDER, filename)
        if not os.path.exists(doc_path):
            return jsonify({'error': '文件不存在'}), 404

        # 确保输出文件名有.xlsx扩展名
        if not output_filename.endswith('.xlsx'):
            if output_filename:
                output_filename = output_filename + '.xlsx'
            else:
                output_filename = 'test_cases.xlsx'

        # 构建输出路径
        output_path = os.path.join(os.path.dirname(UPLOAD_FOLDER), output_filename)

        # 构建AI生成系统的命令
        ai_main_path = os.path.join(os.path.dirname(UPLOAD_FOLDER), 'src', 'main.py')
        template_path = os.path.join(os.path.dirname(UPLOAD_FOLDER), 'src', 'templates',
                                     f'{test_type}_test_template.json')

        # 使用绝对路径，确保路径正确
        python_executable = os.path.abspath(sys.executable)
        ai_main_path_abs = os.path.abspath(ai_main_path)
        doc_path_abs = os.path.abspath(doc_path)
        output_path_abs = os.path.abspath(output_path)

        # 验证所有路径
        if not os.path.exists(python_executable):
            return jsonify({'error': f'Python解释器不存在: {python_executable}'}), 500
        
        if not os.path.exists(ai_main_path_abs):
            return jsonify({'error': f'AI主程序不存在: {ai_main_path_abs}'}), 500
        
        if not os.path.exists(doc_path_abs):
            return jsonify({'error': f'文档文件不存在: {doc_path_abs}'}), 500

        # 确保输出目录存在
        output_dir = os.path.dirname(output_path_abs)
        os.makedirs(output_dir, exist_ok=True)

        # 构建跨平台兼容的命令参数列表
        cmd_args = [
            python_executable,
            ai_main_path_abs,
            '-d', doc_path_abs,
            '-t', test_type,
            '-o', output_path_abs,
            '-c', str(concurrency)
        ]

        # 添加调试日志
        print(f"执行AI生成系统命令:")
        print(f"Python解释器: {python_executable}")
        print(f"AI主程序: {ai_main_path_abs}")
        print(f"文档路径: {doc_path_abs}")
        print(f"输出路径: {output_path_abs}")
        print(f"测试类型: {test_type}")
        print(f"并发数: {concurrency}")
        print(f"工作目录: {os.path.dirname(UPLOAD_FOLDER)}")
        print(f"完整命令: {' '.join(cmd_args)}")

        # 执行AI生成系统
        try:
            # 检查AI主程序文件是否存在和可执行
            if not os.path.exists(ai_main_path_abs):
                return jsonify({'error': f'AI主程序文件不存在: {ai_main_path_abs}'}), 500
            
            # 在Unix系统上检查文件权限
            if os.name != 'nt':  # 不是Windows系统
                if not os.access(ai_main_path_abs, os.R_OK):
                    return jsonify({'error': f'AI主程序文件无读取权限: {ai_main_path_abs}'}), 500

            result = subprocess.run(
                cmd_args,
                capture_output=True,
                text=True,
                encoding='utf-8',  # 明确指定UTF-8编码
                errors='replace',  # 遇到无法解码的字符时替换为占位符
                shell=False,  # 不使用shell，直接传递参数列表
                cwd=os.path.dirname(UPLOAD_FOLDER),  # 设置工作目录
                timeout=1800  # 30分钟超时，因为AI系统需要很长时间处理多层过滤
            )

            if result.returncode == 0:
                # 检查输出文件是否存在
                if os.path.exists(output_path):
                    return jsonify({
                        'success': True,
                        'message': '测试用例生成成功',
                        'output_file': output_filename,
                        'output_path': output_path
                    })
                else:
                    return jsonify({
                        'success': False,
                        'error': '生成成功但输出文件未找到',
                        'stdout': result.stdout,
                        'stderr': result.stderr
                    }), 500
            else:
                return jsonify({
                    'success': False,
                    'error': '测试用例生成失败',
                    'stdout': result.stdout,
                    'stderr': result.stderr,
                    'returncode': result.returncode,
                    'platform': os.name,
                    'command': ' '.join(cmd_args)
                }), 500

        except subprocess.TimeoutExpired:
            return jsonify({'error': '生成超时（30分钟），AI系统需要更长时间处理复杂的多层过滤逻辑，请稍后重试'}), 408
        except FileNotFoundError as e:
            return jsonify({'error': f'文件未找到: {str(e)}', 'platform': os.name, 'command': ' '.join(cmd_args)}), 500
        except PermissionError as e:
            return jsonify({'error': f'权限错误: {str(e)}', 'platform': os.name, 'command': ' '.join(cmd_args)}), 500
        except Exception as e:
            return jsonify({'error': f'执行失败: {str(e)}', 'platform': os.name, 'command': ' '.join(cmd_args)}), 500

    except Exception as e:
        return jsonify({'error': f'生成失败: {str(e)}'}), 500


@ai_generate_bp.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    """下载生成的测试用例文件"""
    try:
        # 构建文件路径
        file_path = os.path.join(os.path.dirname(UPLOAD_FOLDER), filename)

        if not os.path.exists(file_path):
            return jsonify({'error': '文件不存在'}), 404

        # 发送文件
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': f'下载失败: {str(e)}'}), 500


@ai_generate_bp.route('/delete/<filename>', methods=['DELETE'])
def delete_file(filename):
    """删除生成的测试用例文件"""
    try:
        # 构建文件路径
        file_path = os.path.join(os.path.dirname(UPLOAD_FOLDER), filename)

        if not os.path.exists(file_path):
            return jsonify({'error': '文件不存在'}), 404

        # 检查文件扩展名，只允许删除Excel文件
        if not filename.endswith('.xlsx'):
            return jsonify({'error': '只能删除Excel文件'}), 400

        # 删除文件
        os.remove(file_path)

        return jsonify({
            'success': True,
            'message': f'文件 {filename} 删除成功'
        })

    except PermissionError:
        return jsonify({'error': '没有权限删除文件'}), 403
    except Exception as e:
        return jsonify({'error': f'删除失败: {str(e)}'}), 500


@ai_generate_bp.route('/files', methods=['GET'])
def list_generated_files():
    """列出已生成的文件"""
    try:
        ai_test_cases_dir = os.path.dirname(UPLOAD_FOLDER)
        files = []

        for filename in os.listdir(ai_test_cases_dir):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(ai_test_cases_dir, filename)
                file_stat = os.stat(file_path)
                files.append({
                    'filename': filename,
                    'size': file_stat.st_size,
                    'created_at': datetime.fromtimestamp(file_stat.st_ctime).isoformat(),
                    'modified_at': datetime.fromtimestamp(file_stat.st_mtime).isoformat()
                })

        # 按修改时间排序，最新的在前
        files.sort(key=lambda x: x['modified_at'], reverse=True)

        return jsonify({
            'success': True,
            'files': files
        })

    except Exception as e:
        return jsonify({'error': f'获取文件列表失败: {str(e)}'}), 500


# 新增：从生成的Excel统计汇总信息
@ai_generate_bp.route('/summary', methods=['GET'])
def get_generated_file_summary():
    """读取生成的用例Excel，统计优先级分布、总数、各Category数量及文件属性"""
    try:
        from openpyxl import load_workbook

        filename = request.args.get('file')
        test_type = request.args.get('test_type', 'functional')  # 获取用户选择的测试类型
        if not filename:
            return jsonify({'error': '缺少参数: file'}), 400

        # 组装文件路径（生成文件保存于 ai_test_cases 目录）
        ai_test_cases_dir = os.path.dirname(UPLOAD_FOLDER)
        file_path = os.path.join(ai_test_cases_dir, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': '文件不存在'}), 404

        # 打开Excel（只读，读计算结果）
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active

        # 读取表头，建立列索引
        header = None
        header_row_idx = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=1, values_only=True), start=1):
            header = list(row) if row else []
            header_row_idx = r_idx
            break

        if not header:
            # 没有表头或空表
            wb.close()
            file_stat = os.stat(file_path)
            created_at = datetime.fromtimestamp(file_stat.st_ctime).isoformat()
            modified_at = datetime.fromtimestamp(file_stat.st_mtime).isoformat()
            return jsonify({
                'success': True,
                'summary': {
                    'total_cases': 0,
                    'priority_counts': {},
                    'category_counts': {},
                    'file_size': file_stat.st_size,
                    'created_at': created_at,
                    'modified_at': modified_at
                }
            })

        # 映射列名到索引
        col_index = {str(h).strip(): idx for idx, h in enumerate(header)}
        id_idx = col_index.get('ID')
        priority_idx = col_index.get('Priority')
        category_idx = col_index.get('Category')

        total_cases = 0
        priority_counts = {}
        category_counts = {}

        # 遍历数据行
        for row in ws.iter_rows(min_row=(header_row_idx + 1), values_only=True):
            # 取值
            id_val = row[id_idx] if id_idx is not None and id_idx < len(row) else None
            priority_val = row[priority_idx] if priority_idx is not None and priority_idx < len(row) else None
            category_val = row[category_idx] if category_idx is not None and category_idx < len(row) else None

            # 统计总数（按ID是否有值）
            if id_val is not None and str(id_val).strip() != '':
                total_cases += 1

            # 统计优先级
            if priority_val is not None and str(priority_val).strip() != '':
                key = str(priority_val).strip()
                priority_counts[key] = priority_counts.get(key, 0) + 1

            # 统计Category
            if category_val is not None and str(category_val).strip() != '':
                key = str(category_val).strip()
                category_counts[key] = category_counts.get(key, 0) + 1

        # 文件属性
        file_stat = os.stat(file_path)
        created_at = datetime.fromtimestamp(file_stat.st_ctime).isoformat()
        modified_at = datetime.fromtimestamp(file_stat.st_mtime).isoformat()

        wb.close()

        # 根据用户选择的测试类型，构建要存储到数据库的case_types
        test_type_mapping = {
            'functional': '功能测试',
            'api': '接口测试',
            'ui_auto': 'UI自动化测试'
        }
        selected_test_type = test_type_mapping.get(test_type, '功能测试')

        # 构建要存储到数据库的case_types，只包含用户选择的测试类型
        db_case_types = [selected_test_type]

        summary_data = {
            'total_cases': total_cases,
            'priority_counts': priority_counts,
            'category_counts': db_case_types,  # 返回用户选择的测试类型，而不是Excel中的实际分布
            'functional_test_count': category_counts.get('功能测试', 0),  # 添加真正的功能测试用例数
            'api_test_count': category_counts.get('接口测试', 0),  # 添加真正的接口测试用例数
            'ui_auto_test_count': category_counts.get('UI自动化测试', 0),  # 添加真正的UI自动化测试用例数
            'file_size': file_stat.st_size,
            'created_at': created_at,
            'modified_at': modified_at
        }

        # 保存汇总数据到数据库（防重复插入）
        try:
            # 先检查是否已存在相同文件名的记录
            existing_record = execute_query("""
                SELECT id FROM ai_test_generation_history 
                WHERE filename = %s
                LIMIT 1
            """, (filename,))

            if existing_record:
                # 如果已存在，则更新现有记录
                execute_query("""
                    UPDATE ai_test_generation_history 
                    SET case_types = %s, priority_distribution = %s, total_cases = %s,
                        functional_test_count = %s, api_test_count = %s, ui_auto_test_count = %s,
                        estimated_file_size = %s, generated_at = %s
                    WHERE filename = %s
                """, (
                    json.dumps(db_case_types),
                    json.dumps(priority_counts),
                    total_cases,
                    category_counts.get('功能测试', 0),
                    category_counts.get('接口测试', 0),
                    category_counts.get('UI自动化测试', 0),
                    file_stat.st_size,
                    datetime.now(),
                    filename
                ), fetch=False)
                print(f"更新现有记录: {filename}")
            else:
                # 如果不存在，则插入新记录
                execute_query("""
                    INSERT INTO ai_test_generation_history 
                    (case_types, priority_distribution, total_cases, functional_test_count, 
                     api_test_count, ui_auto_test_count, estimated_file_size, generated_at, filename)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    json.dumps(db_case_types),
                    json.dumps(priority_counts),
                    total_cases,
                    category_counts.get('功能测试', 0),
                    category_counts.get('接口测试', 0),
                    category_counts.get('UI自动化测试', 0),
                    file_stat.st_size,
                    datetime.now(),
                    filename
                ), fetch=False)
                print(f"插入新记录: {filename}")

        except Exception as db_error:
            print(f"保存汇总数据到数据库失败: {str(db_error)}")

        return jsonify({
            'success': True,
            'summary': summary_data
        })
    except Exception as e:
        return jsonify({'error': f'获取汇总失败: {str(e)}'}), 500


@ai_generate_bp.route('/latest_summary', methods=['GET'])
def get_latest_generation_summary():
    """获取最新一次AI生成用例的汇总信息"""
    try:
        result = execute_query("""
            SELECT case_types, priority_distribution, total_cases, 
                   functional_test_count, api_test_count, ui_auto_test_count,
                   estimated_file_size, generated_at, filename
            FROM ai_test_generation_history 
            ORDER BY generated_at DESC 
            LIMIT 1
        """)

        if not result:
            return jsonify({
                'success': True,
                'summary': None
            })

        # 获取第一条记录
        record = result[0]

        # 解析JSON字段
        case_types = json.loads(record['case_types']) if record['case_types'] else {}
        priority_distribution = json.loads(record['priority_distribution']) if record['priority_distribution'] else {}

        summary_data = {
            'total_cases': record['total_cases'],
            'priority_counts': priority_distribution,
            'category_counts': case_types,
            'functional_test_count': record['functional_test_count'],  # 添加真正的功能测试用例数
            'api_test_count': record['api_test_count'],  # 添加真正的接口测试用例数
            'ui_auto_test_count': record['ui_auto_test_count'],  # 添加真正的UI自动化测试用例数
            'file_size': record['estimated_file_size'],
            'created_at': record['generated_at'].isoformat() if record['generated_at'] else None,
            'modified_at': record['generated_at'].isoformat() if record['generated_at'] else None,
            'filename': record['filename']
        }

        return jsonify({
            'success': True,
            'summary': summary_data
        })

    except Exception as e:
        return jsonify({'error': f'获取最新汇总失败: {str(e)}'}), 500
