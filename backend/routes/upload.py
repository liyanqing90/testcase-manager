import os
from flask import Blueprint, request, jsonify, current_app
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import uuid
from datetime import datetime
import logging

# 配置日志
logger = logging.getLogger(__name__)

upload_bp = Blueprint('upload', __name__)
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {'xlsx'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_datetime(dt_str):
    """解析日期时间字符串，如果为空或无效则返回None"""
    if not dt_str or dt_str.strip() == "":
        return None
    try:
        # 尝试解析日期时间
        return datetime.strptime(dt_str.strip(), "%Y-%m-%d %H:%M:%S")
    except ValueError:
        try:
            # 尝试解析日期
            return datetime.strptime(dt_str.strip(), "%Y-%m-%d")
        except ValueError:
            # 如果都失败了，返回None
            return None

@upload_bp.route('/upload_case', methods=['POST'])
def upload_case():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'File type not allowed'}), 400
    
    # 获取项目ID参数
    project_id = request.form.get('project_id')
    
    # 改进文件名处理逻辑，确保中文文件名能被正确处理
    filename = secure_filename(file.filename)
    # 如果secure_filename处理后文件名不正确，生成一个唯一的文件名
    if filename == '' or '.' not in filename:
        filename = str(uuid.uuid4()) + '.xlsx'
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    # 解析 Excel，适配用户表头和内容
    try:
        wb = load_workbook(filepath)
    except InvalidFileException:
        # 删除无效文件
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'error': '请上传有效的 .xlsx 文件'}), 400
    except Exception as e:
        # 删除无效文件
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'error': '文件解析错误: ' + str(e)}), 400
    
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        case = {}
        for h, v in zip(headers, row):
            if v is None or (isinstance(v, str) and v.strip() == "空"):
                case[h] = ""
            else:
                case[h] = str(v).strip() if v is not None else ""
        cases.append(case)
    
    # 检查文件内重复ID（不区分项目）
    id_counts = {}
    for case in cases:
        case_id = case.get('ID')
        if case_id:
            id_counts[case_id] = id_counts.get(case_id, 0) + 1
    
    # 标记文件内重复的用例
    for case in cases:
        case_id = case.get('ID')
        if case_id and id_counts[case_id] > 1:
            case['duplicate'] = True
            case['duplicate_reason'] = '文件内重复'
        else:
            case['duplicate'] = False
            case['duplicate_reason'] = ''
    
    # 同时查询数据库中的重复项（全局唯一性检查）
    import sys
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from backend.app import mysql
    cur = mysql.connection.cursor()
    case_ids = [c.get('ID') for c in cases if c.get('ID')]
    if case_ids:
        format_strings = ','.join(['%s'] * len(case_ids))
        # 修改查询逻辑，只检查当前项目中的重复用例，而不是整个数据库
        cur.execute(f"SELECT case_id FROM test_cases WHERE case_id IN ({format_strings}) AND project_id = %s", 
                   tuple(case_ids + [project_id]))
        existing = set(row['case_id'] for row in cur.fetchall())
        # 更新重复标记，添加数据库重复的信息
        for case in cases:
            if case.get('ID') in existing and not case['duplicate']:
                case['duplicate'] = True
                case['duplicate_reason'] = '数据库中存在'
    else:
        existing = set()
    
    return jsonify({'cases': cases, 'headers': headers})

@upload_bp.route('/import_case', methods=['POST'])
def import_case():
    data = request.json
    cases = data.get('cases', [])
    project_id = data.get('project_id')  # 获取项目ID
    # 延迟导入app模块并获取mysql实例
    import sys
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from backend.app import mysql
    cur = mysql.connection.cursor()
    imported = []
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    for c in cases:
        # 支持新ID映射
        case_id = c.get('new_ID') or c.get('ID')
        
        # 验证case_id长度，如果超过200字符则截断
        if case_id and len(str(case_id)) > 200:
            original_case_id = case_id
            case_id = str(case_id)[:200]  # 截断到200字符
            logger.warning(f"用例ID过长，已截断: {original_case_id[:50]}... -> {case_id}")
        
        # 处理日期时间字段
        created_at = c.get('Created At')
        updated_at = c.get('Updated At')
        
        # 如果日期时间字段为空，则使用当前时间
        if not created_at or created_at.strip() == "":
            created_at = current_time
        if not updated_at or updated_at.strip() == "":
            updated_at = current_time
        
        # 处理Created By字段，默认为"导入"
        created_by = c.get('Created By')
        if not created_by or created_by.strip() == "":
            created_by = "导入"
        
        try:
            # 插入测试用例
            cur.execute(
                """
                INSERT INTO test_cases
                (case_id, title, description, preconditions, steps, expected_results, priority, category, status, created_at, updated_at, created_by, last_updated_by, project_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    case_id,
                    c.get('Title') or "",
                    c.get('Description') or "",
                    c.get('Preconditions') or "",
                    c.get('Steps') or "",
                    c.get('Expected Results') or "",
                    c.get('Priority') or "",
                    c.get('Category') or "",
                    c.get('Status') or "",
                    created_at,
                    updated_at,
                    created_by,  # 使用处理后的created_by值
                    c.get('Last Updated By') or "",
                    project_id  # 添加项目ID
                )
            )
            
            # 获取插入的测试用例ID
            test_case_id = cur.lastrowid
            
            # 同时在project_cases表中创建关联记录
            cur.execute(
                "INSERT INTO project_cases (project_id, test_case_id) VALUES (%s, %s)",
                (project_id, test_case_id)
            )
            
            imported.append(case_id)
        except Exception as e:
            mysql.connection.rollback()
            error_msg = str(e)
            
            # 检查是否是重复键错误
            if "Duplicate entry" in error_msg and "unique_case_id_per_project" in error_msg:
                # 提取重复的case_id
                import re
                match = re.search(r"Duplicate entry '[^']+-([^']+)' for key 'unique_case_id_per_project'", error_msg)
                if match:
                    duplicate_case_id = match.group(1)
                    return jsonify({
                        'error': f'用例ID "{duplicate_case_id}" 在当前项目中已存在，请使用不同的ID',
                        'failed_case': duplicate_case_id,
                        'error_type': 'duplicate_case_id'
                    }), 400
                else:
                    return jsonify({
                        'error': '用例ID在当前项目中已存在，请使用不同的ID',
                        'failed_case': case_id,
                        'error_type': 'duplicate_case_id'
                    }), 400
            else:
                return jsonify({'error': error_msg, 'failed_case': case_id}), 400
    mysql.connection.commit()
    return jsonify({'imported': imported, 'count': len(imported)})